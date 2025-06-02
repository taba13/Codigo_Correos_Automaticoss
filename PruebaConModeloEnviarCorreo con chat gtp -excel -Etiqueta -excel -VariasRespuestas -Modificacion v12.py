import requests # type: ignore
import time
from bs4 import BeautifulSoup # type: ignore
from msal import PublicClientApplication # type: ignore
from datetime import datetime, timedelta,timezone
from openpyxl import load_workbook
from unidecode import unidecode
import json
import os
import re
from openai import AsyncOpenAI
import asyncio
import pandas as pd
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import threading
import matplotlib.pyplot as plt

# Configuración del cliente de OpenAI (Asincrónico)
client = AsyncOpenAI(api_key=os.getenv('OPENAI_API_KEY'))

# Archivos para almacenar los IDs de correos e hilos procesados
CORREOS_PROCESADOS_FILE = 'correos_procesados.json'
HILOS_PROCESADOS_FILE = 'hilos_procesados.json'

# Lista para almacenar los correos procesados que se exportarán a Excel
correos_exportar = []

# Clase para interactuar con Microsoft Graph API
class GraphClient:
    def __init__(self, client_id, authority):
        self.client_id = client_id
        self.authority = authority
        self.scope = ['https://graph.microsoft.com/.default']
        self.app = PublicClientApplication(self.client_id, authority=self.authority)
        self.access_token = None
        self.token_expires_at = None

    def authenticate(self, username, password):
        try:
            result = self.app.acquire_token_by_username_password(username, password, scopes=self.scope)
            if "access_token" in result:
                self.access_token = result['access_token']
                self.token_expires_at = datetime.now(timezone.utc) + timedelta(hours=1)
                print(f"Token de acceso delegado obtenido exitosamente para {username}.")
            else:
                print(f"Error al obtener el token de acceso delegado para {username}.")
                print(result.get("error"))
                print(result.get("error_description"))
                self.access_token = None
        except Exception as e:
            print(f"Error durante la autenticación para {username}: {e}")
            self.access_token = None

    def get_access_token(self, username, password):
        if self.access_token is None or self.token_expires_at is None or self.token_expires_at < datetime.utcnow():
            self.authenticate(username, password)
        return self.access_token

def cargar_ids_desde_archivo(archivo):
    if os.path.exists(archivo):
        if os.stat(archivo).st_size > 0:  # Verificar si el archivo no está vacío
            with open(archivo, 'r') as f:
                datos = json.load(f)
                # Si los datos son una lista (formato antiguo), convertirlos a un diccionario
                if isinstance(datos, list):
                    print("El archivo está en formato de lista. Convirtiendo a diccionario con fechas.")
                    # Convertimos cada elemento de la lista a un diccionario con la fecha actual
                    return {hilo_id: str(datetime.utcnow()) for hilo_id in datos}
                return datos  # Ya es un diccionario, retornarlo tal cual
        else:
            print(f"El archivo {archivo} está vacío. Retornando un diccionario vacío.")
            return {}
    else:
        return {}



# Guardar correos procesados de manera acumulativa
def guardar_ids_en_archivo(archivo, nuevos_ids):
    ids_existentes = cargar_ids_desde_archivo(archivo)  # Cargar los IDs ya existentes
    ids_existentes.update(nuevos_ids)  # Unir los nuevos con los ya existentes
    with open(archivo, 'w') as f:
     json.dump(ids_existentes, f)  # Guardar el diccionario completo con las fechas

        
def eliminar_hilos_antiguos(archivo, dias=21):
    hilos_procesados = cargar_ids_desde_archivo(archivo)
    if not hilos_procesados:
        return
    
    # Calcular la fecha límite
    fecha_limite = datetime.utcnow() - timedelta(days=dias)

    # Filtrar los hilos procesados que son más recientes que la fecha límite
    hilos_filtrados = {hilo_id: fecha for hilo_id, fecha in hilos_procesados.items() if datetime.strptime(fecha, '%Y-%m-%d %H:%M:%S.%f') > fecha_limite}

    # Guardar los hilos actualizados en el archivo
    with open(archivo, 'w') as f:
        json.dump(hilos_filtrados, f)

    print(f"Se han eliminado los hilos procesados hace más de {dias} días.")

# Cargar correos procesados
def cargar_correos_procesados():
    return cargar_ids_desde_archivo(CORREOS_PROCESADOS_FILE)

# Guardar correos procesados
def guardar_correos_procesados(ids):
    guardar_ids_en_archivo(CORREOS_PROCESADOS_FILE, ids)

# Cargar hilos procesados
def cargar_hilos_procesados():
    return cargar_ids_desde_archivo(HILOS_PROCESADOS_FILE)

def guardar_hilos_procesados(ids):
    eliminar_hilos_antiguos(HILOS_PROCESADOS_FILE)  # Eliminar hilos antiguos antes de guardar
    guardar_ids_en_archivo(HILOS_PROCESADOS_FILE, ids)  # Guardar los hilos con fecha




correos_procesados = cargar_correos_procesados()
hilos_procesados = cargar_hilos_procesados()

archivo_excel = 'respuestas2.xlsx'
hoja_de_respuestas = 'respuestas'
hoja_de_credenciales = 'credenciales'
carpeta_respuestas = 'correos automaticos'

# Cargar credenciales desde Excel
def cargar_credenciales_desde_excel(archivo_excel):
    try:
        wb = load_workbook(archivo_excel)
        sheet = wb[hoja_de_credenciales]
        credenciales = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            email = row[0]
            password = row[1]
            if email and password and email.strip() and password.strip():
                credenciales.append((email, password))
            else:
                print(f"Advertencia: Credenciales incompletas encontradas en la fila {row}.")
        return credenciales
    except Exception as e:
        print(f"Error al cargar credenciales desde el archivo Excel: {str(e)}")
        return None
# Cargar correos a excluir desde una hoja de Excel
def cargar_correos_a_excluir(archivo_excel, hoja_excluir):
    try:
        wb = load_workbook(archivo_excel)
        sheet = wb[hoja_excluir]
        excluir_correos = set()  # Usar un conjunto para buscar más rápido
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Asegurar que la fila tenga datos
                excluir_correos.add(row[0].strip().lower())
        return excluir_correos
    except Exception as e:
        print(f"Error al cargar correos a excluir desde el archivo Excel: {str(e)}")
        return set()

# Cargar datos de respuestas desde Excel incluyendo la columna de preguntas frecuentes
def cargar_datos_desde_excel(archivo_excel, hoja_de_respuestas):
    try:
        wb = load_workbook(archivo_excel)
        sheet = wb[hoja_de_respuestas]
        datos = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 5:
                palabra_clave = row[0]
                pregunta_frecuente = row[2]  # Columna C: Preguntas Frecuentes
                respuesta = row[3]  # Columna D: Respuesta
                sinonimos = row[4].split(',') if row[4] else []
                datos[palabra_clave] = (sinonimos, pregunta_frecuente, respuesta)
        return datos
    except Exception as e:
        print(f"Error al cargar datos desde el archivo Excel: {str(e)}")
        return None


# Cargar etiquetas y preguntas desde el archivo Excel
def cargar_etiquetas_preguntas_desde_excel(archivo_excel, hoja_de_respuestas):
    try:
        wb = load_workbook(archivo_excel)
        sheet = wb[hoja_de_respuestas]
        etiquetas = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            etiqueta = unidecode(row[0].strip().lower())
            pregunta = row[1].strip() if row[1] else ''
            etiquetas.append((etiqueta, pregunta))
        return etiquetas
    except Exception as e:
        print(f"Error al cargar las etiquetas y preguntas desde el archivo Excel: {str(e)}")
        return None

# Monitoreo de cambios en el archivo Excel
class ExcelFileEventHandler(FileSystemEventHandler):
    def __init__(self, archivo_excel, cargar_datos_func, cargar_etiquetas_func):
        self.archivo_excel = archivo_excel
        self.cargar_datos_func = cargar_datos_func
        self.cargar_etiquetas_func = cargar_etiquetas_func

    def on_modified(self, event):
        if event.src_path.endswith(self.archivo_excel):
            print(f"El archivo {self.archivo_excel} ha sido modificado. Recargando datos y etiquetas...")
            # Recargar los datos de la hoja de respuestas y la hoja de etiquetas
            datos = self.cargar_datos_func(self.archivo_excel, hoja_de_respuestas)
            etiquetas = self.cargar_etiquetas_func(self.archivo_excel, hoja_de_respuestas)
            if datos is not None:
                print("Datos de respuestas recargados exitosamente.")
            if etiquetas is not None:
                print("Etiquetas recargadas exitosamente.")
            else:
                print("Error al recargar los datos o etiquetas.")

# Iniciar observador para monitorear cambios en el archivo
def iniciar_observador(archivo_excel, cargar_datos_func, cargar_etiquetas_func):
    event_handler = ExcelFileEventHandler(archivo_excel, cargar_datos_func, cargar_etiquetas_func)
    observer = Observer()
    observer.schedule(event_handler, path=os.path.dirname(os.path.abspath(archivo_excel)), recursive=False)
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

# Obtener ID de carpeta en Microsoft Graph
def obtener_id_carpeta(access_token, carpeta_nombre):
    try:
        url = f"https://graph.microsoft.com/v1.0/me/mailFolders?$filter=displayName eq '{carpeta_nombre}'"
        headers = {'Authorization': 'Bearer ' + access_token}
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            carpeta = response.json().get('value')
            if carpeta:
                carpeta_id = carpeta[0]['id']
                print(f"El ID de la carpeta '{carpeta_nombre}' fue obtenido correctamente.")
                return carpeta_id
            else:
                print(f"La carpeta '{carpeta_nombre}' no existe.")
                return None
        else:
            print(f"Error al obtener la carpeta '{carpeta_nombre}'. Código de estado: {response.status_code}")
            return None
    except Exception as e:
        print(f"Error al obtener la carpeta '{carpeta_nombre}': {str(e)}")
        return None

def obtener_id_carpeta(access_token, carpeta_nombre):
    try:
        url = f"https://graph.microsoft.com/v1.0/me/mailFolders?$filter=displayName eq '{carpeta_nombre}'"
        headers = {'Authorization': 'Bearer ' + access_token}
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            carpeta = response.json().get('value')
            if carpeta:
                carpeta_id = carpeta[0]['id']
                print(f"El ID de la carpeta '{carpeta_nombre}' fue obtenido correctamente.")
                return carpeta_id
            else:
                print(f"La carpeta '{carpeta_nombre}' no existe.")
                return None
        else:
            print(f"Error al obtener la carpeta '{carpeta_nombre}'. Código de estado: {response.status_code}")
            return None
    except Exception as e:
        print(f"Error al obtener la carpeta '{carpeta_nombre}': {str(e)}")
        return None

async def obtener_etiqueta_openai(texto, etiquetas):
    try:
        # Crear la lista de categorías disponibles
        categorias = etiquetas
        print("Categorías disponibles:", categorias)

        # Primera consulta a OpenAI con las etiquetas predefinidas
        sistema = f"""
               Eres un asistente encargado de analizar mensajes y clasificar su contenido 
               
               Instrucciones:
               1. Evalúa el mensaje completo y determina:
                  - **Necesidad principal**: Identifica la necesidad más importante del mensaje.
                  - **Subnecesidades** (opcional): Si el mensaje incluye detalles adicionales relevantes, clasifícalos como subnecesidades relacionadas.
               2. No bases tu análisis únicamente en palabras clave. Considera el significado completo del mensaje, el tono, y el propósito general para determinar la categoría correcta.
               3. Si ninguna de las categorías disponibles aplica al mensaje, responde con: "sin etiqueta".
               4. Devuelve únicamente las etiquetas como respuesta, separadas por comas. 
                  - Si hay una necesidad principal y subnecesidades, organiza las etiquetas con la principal primero, seguida de las subnecesidades.
               5 - Depues de obtener la necesidades claras listalas y devuelvelas y comparalas con estos parametros o requesitos {categorias} despues de compararla
                   se debe tener el 80% de similutud para dar la categoria 


               
               Formato de Respuesta:
               - Si solo hay una necesidad: "categoria_principal".
               - Si hay una necesidad principal y subnecesidades: "categoria_principal, subcategoria_1, subcategoria_2".

        """

        # Solicitud a OpenAI para la primera clasificación
        completion = await client.chat.completions.create(
            model='gpt-3.5-turbo',
            messages=[
                {"role": "system", "content": sistema},
                {"role": "user", "content": f"Mensaje: {texto}"}
            ],
            max_tokens=200
        )

        # Procesar la respuesta de OpenAI
        respuesta = completion.choices[0].message.content.strip().lower()

        # Verificar si la respuesta es "sin etiqueta"
        if respuesta == "sin etiqueta":
            print("El mensaje no corresponde a ninguna categoría. Intentando identificar la necesidad real del remitente...")

            # Segunda instrucción para identificar la necesidad real sin depender de las categorías predefinidas
            sistema_refinado = """
            Eres un asistente que analiza correos electrónicos para identificar la necesidad real del remitente.
            Tu tarea es leer el mensaje completo y devolver una etiqueta eque describa la necesidad principal , incluso si no está en una lista predefinida.
            Si no puedes identificar ninguna necesidad, responde con la nueva etiqueta definida en dos palabras solo la etiqueta sin justificacion.
            """

            # Segunda solicitud a OpenAI para identificar la necesidad real
            completion_refinado = await client.chat.completions.create(
                model='gpt-3.5-turbo',
                messages=[
                    {"role": "system", "content": sistema_refinado},
                    {"role": "user", "content": f"Mensaje: {texto}"}
                ],
                max_tokens=200
            )

            # Procesar la segunda respuesta
            respuesta_refinada = completion_refinado.choices[0].message.content.strip().lower()

            if respuesta_refinada == "sin etiqueta":
                print("El mensaje no corresponde a ninguna necesidad identificada incluso tras la segunda consulta.")
                return None

            etiquetas_asignadas = [etiqueta.strip() for etiqueta in respuesta_refinada.split(',')]
            print(f"Etiquetas asignadas por OpenAI (consulta refinada): {etiquetas_asignadas}")
            return etiquetas_asignadas

        # Si la primera consulta devuelve etiquetas válidas, retornarlas
        etiquetas_asignadas = [etiqueta.strip() for etiqueta in respuesta.split(',')]
        print(f"Etiquetas asignadas por OpenAI: {etiquetas_asignadas}")
        return etiquetas_asignadas

    except Exception as e:
        print(f"Error al obtener la etiqueta desde OpenAI: {str(e)}")
        return None




# Obtener el último correo de un hilo de conversación
def obtener_ultimo_correo_hilo(conversation_id, access_token):
    correos = obtener_correos_hilo(conversation_id, access_token)
    if correos:
        correos.sort(key=lambda x: x['receivedDateTime'], reverse=True)
        return correos[0]
    return None

# Obtener correos de un hilo de conversación
def obtener_correos_hilo(conversation_id, access_token):
    try:
        url = f"https://graph.microsoft.com/v1.0/me/messages?$filter=conversationId eq '{conversation_id}'"
        headers = {'Authorization': 'Bearer ' + access_token}
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            correos = response.json().get('value', [])
            return correos
        else:
            print(f"Error al obtener correos del hilo. Código de estado: {response.status_code}")
            return None
    except Exception as e:
        print(f"Error al obtener correos del hilo: {str(e)}")
        return None


# Modificar la función de procesamiento de correos para excluir los no deseados
async def procesar_correos(access_token, carpeta_respuestas_id, datos, etiquetas, current_email):
    start_date = (datetime.utcnow() - timedelta(days=1)).isoformat() + 'Z'
    end_date = datetime.utcnow().isoformat() + 'Z'
    correos = obtener_correos_nuevos(access_token, start_date, end_date)
    hilos_procesados_local = {}

    # Cargar la lista de exclusión
    correos_a_excluir = cargar_correos_a_excluir(archivo_excel, "excluir")
    print(f"Se cargaron {len(correos_a_excluir)} correos para excluir.")

    tareas = []
    for correo in correos:
        remitente = correo.get('sender', {}).get('emailAddress', {}).get('address', '').lower()

        # Excluir si el correo está en la lista
        if remitente in correos_a_excluir:
            print(f"El correo de {remitente} está en la lista de exclusión. Se omitirá.")
            continue

        hilo_id = correo.get('conversationId', '')
        if hilo_id in hilos_procesados or hilo_id in hilos_procesados_local:
            print(f"El hilo {hilo_id} ya fue procesado. Se omitirá.")
            continue

        ultimo_correo = obtener_ultimo_correo_hilo(hilo_id, access_token)
        if ultimo_correo:
            tareas.append(procesar_mensaje(ultimo_correo, access_token, carpeta_respuestas_id, datos, etiquetas, current_email))
            hilos_procesados_local[hilo_id] = str(datetime.utcnow())

    await asyncio.gather(*tareas)
    guardar_hilos_procesados(hilos_procesados_local)

def obtener_correos_nuevos(access_token, start_date, end_date, top=1000):
    try:
        url = f"https://graph.microsoft.com/v1.0/me/mailFolders/inbox/messages"
        headers = {'Authorization': 'Bearer ' + access_token}
        params = {
            '$filter': f"receivedDateTime ge {start_date} and receivedDateTime le {end_date}",
            '$select': 'sender,subject,body,receivedDateTime,id,isRead,conversationId',
            '$top': top  # Aquí definimos cuántos correos queremos recibir
        }
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            correos = response.json().get('value', [])
            print(f"Se encontraron {len(correos)} correos nuevos.")
            return correos
        else:
            print(f"Error al obtener correos nuevos. Código de estado: {response.status_code}")
            return []
    except Exception as e:
        print(f"Error al obtener correos nuevos: {str(e)}")
        return []



# Responder al correo en formato HTML para permitir el uso de etiquetas como <strong> y preservar espacios y saltos de línea
def responder_correo(correo_id, access_token, respuesta):
    try:
        mensaje_adicional = """
        <br><br>_______________________________________________________<br>
        Desde la dirección administrativa y financiera le saludamos<br><br>
        Esto es una respuesta automática<br><br>
        Si esta información no corresponde a su pregunta, por favor responda este correo con la palabra 'Revisar'.
        """

        # Convertir saltos de línea a <br> y espacios a &nbsp; en la respuesta
        respuesta_html = respuesta.replace('\n', '<br>').replace('  ', '&nbsp;&nbsp;')

        url = f"https://graph.microsoft.com/v1.0/me/messages/{correo_id}/reply"
        headers = {'Authorization': 'Bearer ' + access_token, 'Content-Type': 'application/json'}
        payload = {
            "message": {
                "body": {
                    "contentType": "HTML",  # Aseguramos que el contenido se envíe como HTML
                    "content": respuesta_html + mensaje_adicional
                }
            }
        }
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 202:
            print(f"Correo {correo_id} respondido exitosamente.")
        else:
            print(f"No se pudo responder el correo {correo_id}.")
    except Exception as e:
        print(f"Error al responder el correo {correo_id}: {str(e)}")



# Mover correo a carpeta específica
def mover_correo_a_carpeta(correo_id, access_token, carpeta_id, tipo_carpeta):
    try:
        url = f"https://graph.microsoft.com/v1.0/me/messages/{correo_id}/move"
        headers = {'Authorization': 'Bearer ' + access_token, 'Content-Type': 'application/json'}
        payload = {"destinationId": carpeta_id}
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 201:
            print(f"Correo {correo_id} movido a la carpeta de {tipo_carpeta} exitosamente.")
        else:
            print(f"No se pudo mover el correo {correo_id} a la carpeta de {tipo_carpeta}.")
    except Exception as e:
        print(f"Error al mover el correo {correo_id} a la carpeta de {tipo_carpeta}: {str(e)}")

# Texto del aviso de confidencialidad que debe eliminarse
AVISO_CONFIDENCIALIDAD = """
Aviso de confidencialidad
Este mensaje es confidencial, puede contener informacion privilegiada y no puede ser usado ni divulgado por personas distintas de su destinatario. 
Si obtiene esta transmision por error, por favor destruya su contenido y avise a su remitente. 
Esta prohibida su retencion, grabacion, utilizacion, aprovechamiento o divulgacion con cualquier proposito. 
Este mensaje ha sido sometido a programas antivirus. No obstante, la Universidad de Manizales no asume ninguna responsabilidad por eventuales danos generados 
por el recibo y el uso de este material, siendo responsabilidad del destinatario verificar con sus propios medios la existencia de defectos, virus u otros riesgos de seguridad. 
El presente correo electronico solo refleja la opinion de su remitente y no representa necesariamente la opinion oficial de la Universidad de Manizales. 
Aprovecha los medios digitales para comunicarte. Antes de imprimir, revisa que tan necesario es. Gracias por tu aporte en la construccion de Universidad Sostenible...
"""

# Función para eliminar el aviso de confidencialidad del contenido del correo
def limpiar_aviso_confidencialidad(texto):
    return texto.replace(AVISO_CONFIDENCIALIDAD.strip(), '').strip()

# Procesar mensajes y asignar respuestas, considerando la columna de preguntas frecuentes
# Función para convertir URLs de texto en enlaces clicables
def convertir_enlaces_clicables(texto):
    # Expresión regular para encontrar URLs en el texto
    url_regex = r'(https?://\S+)'
    # Reemplaza las URLs encontradas por la versión clicable en HTML
    return re.sub(url_regex, r'<a href="\1" target="_blank">\1</a>', texto)

# Procesar mensajes y asignar respuestas, además de almacenar datos en correos_exportar
async def procesar_mensaje(ultimo_correo, access_token, carpeta_respuestas_id, datos, etiquetas, current_email):
    try:
        correo_id = ultimo_correo['id']
        hilo_id = ultimo_correo.get('conversationId', '')

        # Verificar si el correo o hilo ya fue procesado
        if correo_id in correos_procesados or hilo_id in hilos_procesados:
            print("El correo o el hilo ya fue procesado. Se omitirá.")
            return

        remitente = ultimo_correo.get('sender', {}).get('emailAddress', {}).get('address', '').lower()
        if remitente == current_email:
            print(f"El correo es del remitente actual ({current_email}). No se procesará.")
            return

        # Verificar si el correo tiene archivos adjuntos
        if ultimo_correo.get('hasAttachments', False):
            print(f"El correo {correo_id} tiene archivos adjuntos. Se omitirá.")
            return

        # Limpiar el aviso de confidencialidad del cuerpo del mensaje
        cuerpo_correo = ultimo_correo.get('body', {}).get('content', '')
        cuerpo_limpio = limpiar_aviso_confidencialidad(cuerpo_correo)

        # Buscar etiqueta asignada por OpenAI usando el cuerpo del mensaje limpio
        etiquetas_asignadas = await obtener_etiqueta_openai(cuerpo_limpio, etiquetas)

        # Actualizar el diccionario de correos y hilos procesados
        correos_procesados[correo_id] = str(datetime.utcnow())
        hilos_procesados[hilo_id] = str(datetime.utcnow())

        guardar_correos_procesados(correos_procesados)
        guardar_hilos_procesados(hilos_procesados)

        if etiquetas_asignadas is None:
            print("No se encontró ninguna etiqueta adecuada.")
            etiquetas_asignadas = ['sin etiqueta']

        respuesta_completa = []
        for etiqueta in etiquetas_asignadas:
            etiqueta_normalizada = unidecode(etiqueta.strip().lower())  # Normalizar etiqueta para comparación
            if etiqueta_normalizada in datos:
                _, pregunta_frecuente, respuesta_etiqueta = datos[etiqueta_normalizada]
                # Convertir enlaces en clicables en la respuesta
                respuesta_etiqueta_clicable = convertir_enlaces_clicables(respuesta_etiqueta)
                # Resaltar el título usando HTML para que aparezca en negrita
                respuesta_completa.append(f"<strong>{pregunta_frecuente}</strong>\n{respuesta_etiqueta_clicable}")

        respuesta_final = '\n\n'.join(respuesta_completa)

        # Extraer el asunto, cuerpo limpio y fecha para exportar
        asunto = unidecode(ultimo_correo.get('subject', ''))
        cuerpo_html = unidecode(cuerpo_limpio)
        fecha = ultimo_correo.get('receivedDateTime', '')

        # Asumiendo que siempre se tendrán dos etiquetas, ajustamos para exportarlas por separado
        etiqueta_primaria = etiquetas_asignadas[0] if len(etiquetas_asignadas) > 0 else ''
        etiqueta_secundaria = etiquetas_asignadas[1] if len(etiquetas_asignadas) > 1 else ''

        # Verificar si el correo ya ha sido exportado previamente antes de agregarlo a la lista de exportación
        if not any(c['correo_id'] == correo_id for c in correos_exportar):
            correos_exportar.append({
                'correo_id': correo_id,  # Agregamos el identificador único del correo
                'asunto': asunto,
                'cuerpo': BeautifulSoup(cuerpo_html, 'html.parser').get_text(separator=' '),
                'fecha': fecha,
                'etiquetas': ', '.join(etiquetas_asignadas),  # Agregar todas las etiquetas separadas por comas
                'respuesta': respuesta_final  # Agregar la respuesta completa con formato
            })

        if respuesta_final:
            responder_correo(correo_id, access_token, respuesta_final) 
            mover_correo_a_carpeta(correo_id, access_token, carpeta_respuestas_id, "respuestas")
        else:
            print(f"No se encontró una respuesta adecuada para el correo {correo_id}.")

    except Exception as e:
        print(f"Error al procesar mensaje: {str(e)}")




def exportar_correos_a_excel():
    try:
        archivo_excel = 'correos_procesados.xlsx'

        # Verificar si el archivo Excel ya existe
        if os.path.exists(archivo_excel):
            # Leer el archivo existente
            df_existente = pd.read_excel(archivo_excel)
            df_nuevo = pd.DataFrame(correos_exportar)

            # Eliminar duplicados basados en la columna 'correo_id' antes de combinar
            df_combined = pd.concat([df_existente, df_nuevo], ignore_index=True).drop_duplicates(subset=['correo_id'])

            # Guardar el archivo Excel con los datos combinados y sin duplicados
            df_combined.to_excel(archivo_excel, index=False)
        else:
            # Crear un nuevo archivo con los correos exportados
            df = pd.DataFrame(correos_exportar)
            df.to_excel(archivo_excel, index=False)

        print("Correos exportados exitosamente a 'correos_procesados.xlsx'.")
    except Exception as e:
        print(f"Error al exportar correos a Excel: {str(e)}")
        
# Archivo temporal para almacenar los IDs de correos ya contados
ARCHIVO_TEMPORAL = 'correos_contados_temp.json'

# Función para cargar los IDs de correos ya contados desde el archivo temporal
def cargar_ids_contados(archivo_temp=ARCHIVO_TEMPORAL):
    if os.path.exists(archivo_temp):
        with open(archivo_temp, 'r') as f:
            return set(json.load(f))
    return set()

# Función para guardar los IDs de correos ya contados en el archivo temporal
def guardar_ids_contados(ids_contados, archivo_temp=ARCHIVO_TEMPORAL):
    with open(archivo_temp, 'w') as f:
        json.dump(list(ids_contados), f)

# Función para contar respuestas y guardar el conteo evitando duplicados y considerando etiquetas individuales
def contar_respuestas_y_guardar_unica(archivo_excel='correos_procesados.xlsx', archivo_conteo='conteo_respuestas_etiquetas.xlsx'):
    try:
        # Cargar los IDs de correos ya contados
        ids_contados = cargar_ids_contados()

        # Leer el archivo Excel con los correos exportados
        if os.path.exists(archivo_excel):
            df = pd.read_excel(archivo_excel)

            # Filtrar los correos que ya han sido contados para evitar duplicados
            df_nuevos = df[~df['correo_id'].isin(ids_contados)]

            # Si no hay correos nuevos, no hacer nada
            if df_nuevos.empty:
                print("No hay nuevos correos para contar.")
                return

            # Separar las etiquetas múltiples y realizar el conteo individualmente
            etiquetas_expandidas = df_nuevos['etiquetas'].str.split(',').explode().str.strip()
            conteo_etiquetas = etiquetas_expandidas.value_counts().reset_index()
            conteo_etiquetas.columns = ['Etiqueta', 'Cantidad de Respuestas']

            # Guardar los nuevos IDs contados en el archivo temporal
            ids_contados.update(df_nuevos['correo_id'].tolist())
            guardar_ids_contados(ids_contados)

            # Guardar el conteo en un archivo Excel para su revisión
            if os.path.exists(archivo_conteo):
                # Si el archivo ya existe, leerlo y actualizarlo con el nuevo conteo
                df_existente = pd.read_excel(archivo_conteo)
                df_actualizado = pd.concat([df_existente, conteo_etiquetas]).groupby('Etiqueta', as_index=False).sum()
                df_actualizado.to_excel(archivo_conteo, index=False)
            else:
                # Crear un nuevo archivo si no existe
                conteo_etiquetas.to_excel(archivo_conteo, index=False)

            print(f"Conteo de respuestas guardado exitosamente en '{archivo_conteo}'.")
        else:
            print(f"El archivo {archivo_excel} no existe.")
    except Exception as e:
        print(f"Error al contar respuestas por etiqueta: {str(e)}")



# Iniciar proceso
async def iniciar_proceso():
    credenciales = cargar_credenciales_desde_excel(archivo_excel)
    if not credenciales:
        print("No se pudieron cargar las credenciales.")
        return

    datos = cargar_datos_desde_excel(archivo_excel, hoja_de_respuestas)
    if datos is None:
        print("No se pudo cargar los datos desde el archivo Excel.")
        return

    etiquetas = cargar_etiquetas_preguntas_desde_excel(archivo_excel, hoja_de_respuestas)
    if not etiquetas:
        print("No se pudieron cargar las etiquetas.")
        return

    for email, password in credenciales:
        print(f"Iniciando procesamiento para {email}...")
        graph_client = GraphClient(client_id='bc650cfd-99fa-4c13-8b1a-f644883c24e0', authority='https://login.microsoftonline.com/4f1e044d-3d70-4990-b26a-b95f0c642e1a')
        access_token = graph_client.get_access_token(email, password)
        if not access_token:
            print(f"No se pudo obtener el token de acceso para {email}.")
            continue

        carpeta_respuestas_id = obtener_id_carpeta(access_token, carpeta_respuestas)
        if not carpeta_respuestas_id:
            print(f"No se pudo obtener el ID de la carpeta de respuestas para {email}.")
            continue

        await procesar_correos(access_token, carpeta_respuestas_id, datos, etiquetas, email)
        print(f"Procesamiento para {email} completado.")
    
    exportar_correos_a_excel()
    contar_respuestas_y_guardar_unica()

# Bucle principal asincrónico
async def iniciar_proceso_y_observador():
    # Iniciar observador en un hilo separado
    observador_thread = threading.Thread(target=iniciar_observador, args=(archivo_excel, cargar_datos_desde_excel, cargar_etiquetas_preguntas_desde_excel))
    observador_thread.daemon = True
    observador_thread.start()

    # Iniciar el proceso principal de procesamiento de correos
    while True:
        await iniciar_proceso()
        await asyncio.sleep(10)

# Ejecutar el bucle de eventos asincrónico
asyncio.run(iniciar_proceso_y_observador())